from fastapi import APIRouter, Depends, Request, Response, Form, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from datetime import date, datetime
from typing import Optional
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.database import get_db
from app.crud import (get_bookings, get_grouped_bookings, update_checkin_day_comments,
                      get_all_services, get_booking_services, add_booking_service,
                      delete_booking_service, get_booking_services_total)
from app.auth import create_session_cookie
from app.config import get_settings
from loguru import logger

router = APIRouter(tags=["web"])
templates = Jinja2Templates(directory="app/templates")
settings = get_settings()


def check_auth(request: Request) -> Optional[str]:
    """Проверка авторизации через session cookie и возврат user_type"""
    session_token = request.cookies.get("session_token")
    if not session_token or session_token != settings.secret_key:
        return None
    
    user_type = request.cookies.get("user_type")
    if not user_type or user_type not in ['admin', 'user']:
        return None
    
    return user_type


@router.get("/", response_class=HTMLResponse)
async def root():
    """Редирект на страницу бронирований"""
    return RedirectResponse(url="/bookings", status_code=302)


@router.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    """Страница входа"""
    return templates.TemplateResponse("login.html", {"request": request})


@router.post("/login")
async def login(
    request: Request,
    response: Response,
    username: str = Form(...),
    password: str = Form(...)
):
    """Обработка формы входа"""
    user_type = None
    if username == settings.admin_username and password == settings.admin_password:
        user_type = 'admin'
    elif username == settings.user_username and password == settings.user_password:
        user_type = 'user'
    
    if user_type:
        response = RedirectResponse(url="/bookings", status_code=302)
        create_session_cookie(response)
        response.set_cookie(
            key="user_type",
            value=user_type,
            httponly=True,
            max_age=86400,  # 24 часа
            samesite="lax"
        )
        logger.info(f"Успешный вход пользователя: {username} ({user_type})")
        return response
    else:
        logger.warning(f"Неудачная попытка входа: {username}")
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "error": "Неверный логин или пароль"},
            status_code=401
        )


@router.get("/logout")
async def logout():
    """Выход из системы"""
    response = RedirectResponse(url="/login", status_code=302)
    response.delete_cookie("session_token")
    response.delete_cookie("user_type")
    logger.info("Пользователь вышел из системы")
    return response


@router.get("/bookings", response_class=HTMLResponse)
async def bookings_page(
    request: Request,
    filter_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Страница со списком бронирований
    Параметры:
    - filter_date: фильтр по дате в формате YYYY-MM-DD (показывает бронирования где дата является датой заселения ИЛИ выселения)
    """
    # Проверка авторизации
    user_type = check_auth(request)
    if not user_type:
        return RedirectResponse(url="/login", status_code=302)
    
    # Преобразуем строку даты в объект date
    date_filter = None
    if filter_date:
        try:
            date_filter = datetime.strptime(filter_date, "%Y-%m-%d").date()
        except ValueError:
            logger.warning(f"Неверный формат даты: {filter_date}")
    
    # Получаем сгруппированные бронирования из БД
    grouped_bookings = get_grouped_bookings(db, filter_date=date_filter)
    
    logger.info(f"Отображение страницы бронирований. Фильтр: {filter_date}, Найдено: {len(grouped_bookings)}")
    
    return templates.TemplateResponse(
        "bookings.html",
        {
            "request": request,
            "grouped_bookings": grouped_bookings,
            "filter_date": filter_date,
            "user_type": user_type
        }
    )


@router.post("/update-checkin-comment")
async def update_checkin_comment(
    request: Request,
    booking_id: int = Form(...),
    comments: str = Form(...),
    db: Session = Depends(get_db)
):
    """
    Обновить комментарии по оплате и проживанию в день заселения
    """
    # Проверка авторизации
    user_type = check_auth(request)
    if not user_type:
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    # Обновляем комментарий
    booking = update_checkin_day_comments(db, booking_id, comments)
    
    if booking:
        logger.info(f"Комментарий обновлен для бронирования {booking_id}")
        return {"status": "success", "message": "Комментарий обновлен"}
    else:
        logger.warning(f"Бронирование {booking_id} не найдено")
        raise HTTPException(status_code=404, detail="Booking not found")


@router.get("/services")
async def get_services_list(
    request: Request,
    db: Session = Depends(get_db)
):
    """
    Получить список всех активных услуг
    """
    # Проверка авторизации
    user_type = check_auth(request)
    if not user_type:
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    services = get_all_services(db, active_only=True)
    return {"services": [{"id": s.id, "name": s.name} for s in services]}


@router.get("/booking-services/{booking_id}")
async def get_booking_services_list(
    request: Request,
    booking_id: int,
    db: Session = Depends(get_db)
):
    """
    Получить список услуг для конкретного бронирования
    """
    # Проверка авторизации
    user_type = check_auth(request)
    if not user_type:
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    services = get_booking_services(db, booking_id)
    total = get_booking_services_total(db, booking_id)
    
    return {
        "services": services,
        "total": total
    }


@router.post("/booking-services")
async def add_service_to_booking(
    request: Request,
    booking_id: int = Form(...),
    service_id: int = Form(...),
    price: float = Form(...),
    db: Session = Depends(get_db)
):
    """
    Добавить услугу к бронированию
    """
    # Проверка авторизации
    user_type = check_auth(request)
    if not user_type:
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    try:
        booking_service = add_booking_service(db, booking_id, service_id, price)
        return {
            "status": "success",
            "message": "Услуга добавлена",
            "id": booking_service.id
        }
    except Exception as e:
        logger.error(f"Ошибка при добавлении услуги: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/booking-services/{booking_service_id}")
async def remove_service_from_booking(
    request: Request,
    booking_service_id: int,
    db: Session = Depends(get_db)
):
    """
    Удалить услугу из бронирования
    """
    # Проверка авторизации
    user_type = check_auth(request)
    if not user_type:
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    success = delete_booking_service(db, booking_service_id)
    
    if success:
        return {"status": "success", "message": "Услуга удалена"}
    else:
        raise HTTPException(status_code=404, detail="Service not found")


@router.get("/export")
async def export_to_excel(
    request: Request,
    filter_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Экспорт таблицы бронирований в Excel
    """
    # Проверка авторизации
    user_type = check_auth(request)
    if not user_type:
        return RedirectResponse(url="/login", status_code=302)
    
    # Преобразуем строку даты в объект date
    date_filter = None
    if filter_date:
        try:
            date_filter = datetime.strptime(filter_date, "%Y-%m-%d").date()
        except ValueError:
            logger.warning(f"Неверный формат даты при экспорте: {filter_date}")
    
    # Получаем сгруппированные бронирования
    grouped_bookings = get_grouped_bookings(db, filter_date=date_filter)
    
    # Создаем Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Бронирования"
    
    # Стили
    header_fill_gray = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    header_fill_red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    header_fill_green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    header_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    thick_border_left = Border(
        left=Side(style='thick'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    thick_border_bottom = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thick')
    )
    thick_border_left_bottom = Border(
        left=Side(style='thick'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thick')
    )
    # Границы для групп дублей
    thick_border_top = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thick'),
        bottom=Side(style='thin')
    )
    thick_border_top_left = Border(
        left=Side(style='thick'),
        right=Side(style='thin'),
        top=Side(style='thick'),
        bottom=Side(style='thin')
    )
    thick_border_both = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )
    thick_border_both_left = Border(
        left=Side(style='thick'),
        right=Side(style='thin'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Заголовки первого уровня (группы)
    # Дата заезда (A1:B1)
    ws.merge_cells('A1:B1')
    date_header = filter_date if filter_date else 'Все даты'
    ws['A1'] = date_header
    ws['A1'].fill = header_fill_gray
    ws['A1'].font = header_font
    ws['A1'].alignment = center_align
    ws['A1'].border = border
    ws['B1'].border = border
    
    # Выселение (C1:E1)
    ws.merge_cells('C1:E1')
    ws['C1'] = 'Выселение'
    ws['C1'].fill = header_fill_gray
    ws['C1'].font = header_font
    ws['C1'].alignment = center_align
    ws['C1'].border = thick_border_left
    for col in ['D', 'E']:
        ws[f'{col}1'].border = border
    
    # Заселение (F1:O1)
    ws.merge_cells('F1:O1')
    ws['F1'] = 'Заселение'
    ws['F1'].fill = header_fill_gray
    ws['F1'].font = header_font
    ws['F1'].alignment = center_align
    ws['F1'].border = thick_border_left
    for col in ['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
        ws[f'{col}1'].border = border
    
    # Заголовки второго уровня (детальные поля)
    headers_row2 = [
        'Адрес', 'Статус дома',
        'ФИО', 'Телефон', 'Комментарий',
        'ФИО', 'Телефон',
        'Дата выселения', 'Кол-во дней', 'Общая сумма', 'Предоплата', 'Доплата', 'Доп. услуги', 'Комментарий', 'Комментарии по оплате и проживанию в день заселения'
    ]
    
    for col_num, header in enumerate(headers_row2, start=1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_align
        
        # Применяем цвет фона в зависимости от колонки
        if col_num <= 2:  # Адрес, Статус дома - серый
            cell.fill = header_fill_gray
            cell.border = thick_border_bottom
        elif col_num <= 5:  # Выселение - бледно-красный
            cell.fill = header_fill_red
            if col_num == 3:  # Первая колонка выселения - жирная граница слева
                cell.border = thick_border_left_bottom
            else:
                cell.border = thick_border_bottom
        else:  # Заселение - бледно-зеленый
            cell.fill = header_fill_green
            if col_num == 6:  # Первая колонка заселения - жирная граница слева
                cell.border = thick_border_left_bottom
            else:
                cell.border = thick_border_bottom
    
    # Данные
    for row_num, group in enumerate(grouped_bookings, start=3):
        # Определяем статус
        if group['checkout'] and group['checkin']:
            status_text = 'Выс/Зас'
        elif group['checkout']:
            status_text = 'Выселение'
        elif group['checkin']:
            status_text = 'Заселение'
        else:
            status_text = ''
        
        # Данные выселения
        checkout_fio = group['checkout'].client_fio if group['checkout'] else ''
        checkout_phone = group['checkout'].client_phone if group['checkout'] else ''
        checkout_notes = group['checkout'].notes if group['checkout'] else ''
        
        # Данные заселения
        checkin_fio = group['checkin'].client_fio if group['checkin'] else ''
        checkin_phone = group['checkin'].client_phone if group['checkin'] else ''
        checkin_end_date = group['checkin'].end_date.strftime('%d.%m.%Y') if group['checkin'] and group['checkin'].end_date else ''
        checkin_nights = group['checkin'].number_of_nights if group['checkin'] else ''
        checkin_amount = float(group['checkin'].amount) if group['checkin'] and group['checkin'].amount else 0
        checkin_prepayment = float(group['checkin'].prepayment) if group['checkin'] and group['checkin'].prepayment else 0
        checkin_doplata = checkin_amount - checkin_prepayment
        checkin_notes = group['checkin'].notes if group['checkin'] else ''
        checkin_day_comments = group['checkin'].checkin_day_comments if group['checkin'] else ''
        
        # Сумма доп. услуг для заселения
        checkin_services_total = 0
        if group['checkin']:
            checkin_services_total = get_booking_services_total(db, group['checkin'].id)
        
        row_data = [
            (group['address'] or '').upper(),
            status_text,
            checkout_fio,
            checkout_phone,
            checkout_notes,
            checkin_fio,
            checkin_phone,
            checkin_end_date,
            checkin_nights,
            checkin_amount,
            checkin_prepayment,
            checkin_doplata,
            checkin_services_total,
            checkin_notes,
            checkin_day_comments
        ]
        
        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(vertical="center")
            
            # Определяем какие границы применять
            has_duplicate = group.get('has_duplicate', False)
            is_first = group.get('is_first_in_group', False)
            is_last = group.get('is_last_in_group', False)
            
            # Жирные границы только для первой/последней строки групп дублей
            has_top_thick = has_duplicate and is_first
            has_bottom_thick = has_duplicate and is_last
            has_left_thick = col_num == 3 or col_num == 6
            
            # Применяем границы в зависимости от позиции и группировки
            if has_top_thick and has_bottom_thick and has_left_thick:
                # Одиночная строка с жирной границей слева
                cell.border = thick_border_both_left
            elif has_top_thick and has_left_thick:
                # Первая строка в группе с жирной границей слева
                cell.border = thick_border_top_left
            elif has_top_thick and has_bottom_thick:
                # Одиночная строка без жирной границы слева
                cell.border = thick_border_both
            elif has_top_thick:
                # Первая строка в группе без жирной границы слева
                cell.border = thick_border_top
            elif has_bottom_thick and has_left_thick:
                # Последняя строка в группе с жирной границей слева
                cell.border = Border(
                    left=Side(style='thick'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thick')
                )
            elif has_bottom_thick:
                # Последняя строка в группе без жирной границы слева
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thick')
                )
            elif has_left_thick:
                # Средняя строка в группе с жирной границей слева
                cell.border = thick_border_left
            else:
                # Средняя строка в группе без жирной границы слева
                cell.border = border
            
            # Делаем адрес жирным (первая колонка)
            if col_num == 1:
                cell.font = Font(bold=True)
            # Делаем "Общая сумма" (10) и "Доплата" (12) жирными
            elif col_num == 10 or col_num == 12:
                cell.font = Font(bold=True)
    
    # Устанавливаем ширину колонок
    column_widths = [25, 15, 20, 18, 25, 15, 15, 15, 12, 15, 15, 15, 15, 35, 35]
    for col_num, width in enumerate(column_widths, start=1):
        column_letter = chr(64 + col_num) if col_num <= 26 else f'A{chr(64 + col_num - 26)}'
        ws.column_dimensions[column_letter].width = width
    
    # Сохраняем в BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Формируем имя файла
    filename = f"bookings_{filter_date if filter_date else 'all'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    logger.info(f"Экспорт {len(grouped_bookings)} бронирований в Excel. Файл: {filename}")
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
