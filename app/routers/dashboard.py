from fastapi import APIRouter, Depends, Request, Response
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from datetime import date, datetime
from typing import Optional, Dict, Any
import re
from functools import lru_cache
import time
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from fastapi.responses import StreamingResponse
from app.database import get_db
from app.auth import create_session_cookie
from app.config import get_settings
from app.crud import get_unique_apartments, get_bookings_by_begin_date, get_payments, get_expenses
from loguru import logger

router = APIRouter(tags=["dashboard"])
templates = Jinja2Templates(directory="app/templates")
settings = get_settings()


def check_auth(request: Request) -> Optional[str]:
    """Проверка авторизации через session cookie и возврат user_type"""
    session_token = request.cookies.get("session_token")
    if not session_token or session_token != settings.secret_key:
        return None
    
    user_type = request.cookies.get("user_type")
    if not user_type or user_type not in ['admin', 'user']:
        return None
    
    # Проверка срока действия для администраторов
    if user_type == 'admin' and settings.admin_expiration_date:
        # Получаем дату создания сессии из cookie (если есть)
        session_created_str = request.cookies.get("session_created")
        if session_created_str:
            try:
                session_created = datetime.fromisoformat(session_created_str)
                # Если сессия была создана ДО даты истечения, то она недействительна
                if session_created.date() < settings.admin_expiration_date.date():
                    return None  # Требуем перелогиниться
                else:
                    # Если сессия была создана ПОСЛЕ даты истечения или в этот же день, то она действительна
                    return user_type
            except ValueError:
                pass
        
        # Если дата создания сессии не указана или не может быть распознана,
        # проверяем текущую дату (для обратной совместимости)
        if datetime.now().date() >= settings.admin_expiration_date.date():
            # Удаляем cookie и возвращаем None для принудительного перелогинивания
            return None
    
    return user_type


def get_base_address(address: str) -> str:
    """Получить базовый адрес без суффикса ДУБЛЬ (регистронезависимо) и без ведущих чисел"""
    if not address:
        return ''
    
    address_clean = address.strip()
    
    # Удаляем ведущие числа в формате '123) ' или '123.4) '
    # Например: "004) 29Б" -> "29Б" или "011.2) Ш15" -> "Ш15"
    address_clean = re.sub(r'^\d+(?:\.\d+)?\)\s*', '', address_clean)
    
    address_upper = address_clean.upper()
    
    # Список суффиксов для удаления (в верхнем регистре)
    suffixes = ['ДУБЛЬ', 'ДУБЛ', 'ДУБЛЕ', 'ДУБ', 'DUBL', 'DOUBLE']
    
    for suffix in suffixes:
        # Проверяем с пробелом перед суффиксом
        suffix_with_space = ' ' + suffix
        if suffix_with_space in address_upper:
            idx = address_upper.find(suffix_with_space)
            result = address_clean[:idx].strip()
            return result
        # Проверяем суффикс в конце без пробела
        elif address_upper.endswith(suffix):
            result = address_clean[:-len(suffix)].strip()
            return result
    
    return address_clean


def get_top_apartments(db: Session, limit: int = 50) -> list:
    """Получить топ N объектов недвижимости для оптимизации"""
    try:
        apartments = get_unique_apartments(db)
        # Ограничиваем количество объектов для обработки
        return apartments[:limit] if len(apartments) > limit else apartments
    except Exception as e:
        logger.error(f"Error in get_top_apartments: {e}")
        return []


def get_monthly_financial_data(db: Session, year: int, month: int, apartments: list) -> dict:
    """Получить финансовые данные за конкретный месяц для ограниченного списка объектов"""
    try:
        start_time = time.time()
        
        # Начало и конец месяца
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year, 12, 31)
        else:
            end_date = date(year, month + 1, 1)
        
        # Структура для хранения данных за месяц
        month_data = {
            'month': month,
            'objects': {},
            'total_income': 0,
            'total_expenses': 0,
            'total_profit': 0,
            'general_expenses': 0  # Добавляем поле для общих расходов
        }
        
        # Для каждого объекта
        for apartment in apartments:
            base_apartment = get_base_address(apartment)
            
            # Получаем бронирования (доходы) для объекта и месяца
            bookings = get_bookings_by_begin_date(
                db, 
                filter_date_from=start_date, 
                filter_date_to=end_date, 
                apartment_title=apartment
            )
            
            # Получаем поступления для объекта и месяца
            payments = get_payments(
                db, 
                filter_date_from=start_date, 
                filter_date_to=end_date, 
                apartment_title=apartment
            )
            
            # Получаем расходы для объекта и месяца
            expenses = get_expenses(
                db, 
                filter_date_from=start_date, 
                filter_date_to=end_date, 
                apartment_title=apartment
            )
            
            # Считаем общую сумму доходов
            income_total = 0
            
            # Доходы от бронирований
            for booking in bookings:
                if booking.amount:
                    income_total += float(booking.amount)
            
            # Доходы от поступлений
            for payment in payments:
                if payment.amount:
                    income_total += float(payment.amount)
            
            # Считаем общую сумму расходов
            expense_total = 0
            for expense in expenses:
                if expense.amount:
                    expense_total += float(expense.amount)
            
            # Прибыль
            profit = income_total - expense_total
            
            # Пропускаем нулевые значения для оптимизации
            if income_total == 0 and expense_total == 0 and profit == 0:
                continue
            
            # Добавляем в структуру данных
            if base_apartment not in month_data['objects']:
                month_data['objects'][base_apartment] = {
                    'apartments': [],
                    'income': 0,
                    'expenses': 0,
                    'profit': 0
                }
            
            # Добавляем объект к группе
            if apartment not in month_data['objects'][base_apartment]['apartments']:
                month_data['objects'][base_apartment]['apartments'].append(apartment)
            
            # Суммируем данные для дублей
            month_data['objects'][base_apartment]['income'] += income_total
            month_data['objects'][base_apartment]['expenses'] += expense_total
            month_data['objects'][base_apartment]['profit'] += profit
            
            # Суммируем общие данные по месяцу
            month_data['total_income'] += income_total
            month_data['total_expenses'] += expense_total
            month_data['total_profit'] += profit
        
        # Получаем общие расходы (не привязанные к объектам)
        all_expenses = get_expenses(db, filter_date_from=start_date, filter_date_to=end_date)
        general_expenses = 0
        for expense in all_expenses:
            # Проверяем, если расход не привязан к конкретному объекту
            if not expense.apartment_title or expense.apartment_title.strip() == '':
                if expense.amount:
                    general_expenses += float(expense.amount)
        
        month_data['general_expenses'] = general_expenses
        
        end_time = time.time()
        logger.debug(f"Processed month {month} in {end_time - start_time:.2f} seconds")
        
        return month_data
    except Exception as e:
        logger.error(f"Error in get_monthly_financial_data: {e}")
        return {}


# Кэшируем результаты на 5 минут
@lru_cache(maxsize=32)
def get_cached_annual_financial_data(cache_key: str, db_hash: int, year: int) -> dict:
    """Кэшированная версия получения финансовых данных за год"""
    # Эта функция будет вызываться с уникальным ключом для кэширования
    # Реальная реализация будет в основном методе
    pass


def get_annual_financial_data(db: Session, year: int) -> dict:
    """Получить финансовые данные за год, сгруппированные по объектам и месяцам"""
    try:
        start_time = time.time()
        
        # Структура для хранения данных
        financial_data = {}
        
        # Получаем ограниченный список объектов
        apartments = get_top_apartments(db)
        
        # Для каждого месяца - получаем данные отдельно
        for month in range(1, 13):
            month_key = f"{year}-{month:02d}"
            financial_data[month_key] = get_monthly_financial_data(db, year, month, apartments)
        
        end_time = time.time()
        logger.info(f"Processed annual financial data for {year} in {end_time - start_time:.2f} seconds")
        
        return financial_data
    except Exception as e:
        logger.error(f"Error in get_annual_financial_data: {e}")
        return {}


def get_apartment_summary(financial_data: dict) -> dict:
    """Получить сводную информацию по объектам за год"""
    try:
        apartment_summary = {}
        
        # Проходим по всем месяцам
        for month_key, month_data in financial_data.items():
            if isinstance(month_data, dict) and 'objects' in month_data:
                for base_apartment, obj_data in month_data['objects'].items():
                    if base_apartment not in apartment_summary:
                        apartment_summary[base_apartment] = {
                            'apartments': obj_data.get('apartments', []),
                            'total_income': 0,
                            'total_expenses': 0,
                            'total_profit': 0
                        }
                    
                    # Суммируем данные за год
                    apartment_summary[base_apartment]['total_income'] += obj_data.get('income', 0)
                    apartment_summary[base_apartment]['total_expenses'] += obj_data.get('expenses', 0)
                    apartment_summary[base_apartment]['total_profit'] += obj_data.get('profit', 0)
        
        return apartment_summary
    except Exception as e:
        logger.error(f"Error in get_apartment_summary: {e}")
        return {}


def calculate_yearly_totals(financial_data: dict) -> dict:
    """Рассчитать итоговые значения за год"""
    try:
        total_income = 0
        total_expenses = 0
        total_profit = 0
        total_general_expenses = 0  # Добавляем учет общих расходов
        
        # Проходим по всем месяцам
        for month_data in financial_data.values():
            if isinstance(month_data, dict):
                total_income += month_data.get('total_income', 0)
                total_expenses += month_data.get('total_expenses', 0)
                total_profit += month_data.get('total_profit', 0)
                total_general_expenses += month_data.get('general_expenses', 0)
        
        return {
            'total_income': total_income,
            'total_expenses': total_expenses,
            'total_profit': total_profit,
            'total_general_expenses': total_general_expenses  # Возвращаем общие расходы
        }
    except Exception as e:
        logger.error(f"Error in calculate_yearly_totals: {e}")
        return {
            'total_income': 0,
            'total_expenses': 0,
            'total_profit': 0,
            'total_general_expenses': 0
        }


# Add custom functions to Jinja2 environment (after they are defined)
templates.env.globals["get_apartment_summary"] = get_apartment_summary
templates.env.globals["calculate_yearly_totals"] = calculate_yearly_totals


@router.get("/dashboard", response_class=HTMLResponse)
async def dashboard_page(
    request: Request,
    year: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """
    Страница дашборда с годовой финансовой информацией
    """
    try:
        start_time = time.time()
        
        # Проверка авторизации (только для админов)
        user_type = check_auth(request)
        if not user_type or user_type != 'admin':
            return Response(status_code=403)
        
        # Если год не указан, используем текущий
        if not year:
            year = datetime.now().year
        
        # Ограничиваем год для оптимизации (не больше 5 лет назад)
        current_year = datetime.now().year
        if year < current_year - 5:
            year = current_year - 5
        elif year > current_year + 5:
            year = current_year + 5
        
        # Получаем финансовые данные
        financial_data = get_annual_financial_data(db, year)
        
        # Вычисляем годовые итоги
        yearly_totals = calculate_yearly_totals(financial_data)
        
        end_time = time.time()
        logger.info(f"Dashboard page loaded for {year} in {end_time - start_time:.2f} seconds")
        
        return templates.TemplateResponse(
            "dashboard.html",
            {
                "request": request,
                "financial_data": financial_data,
                "year": year,
                "user_type": user_type,
                "yearly_totals": yearly_totals  # Добавляем yearly_totals в контекст шаблона
            }
        )
    except Exception as e:
        logger.error(f"Error in dashboard_page: {e}")
        return Response(status_code=500, content="Internal Server Error")


@router.get("/dashboard/export")
async def export_dashboard_to_excel(
    request: Request,
    year: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """
    Экспорт дашборда в Excel
    """
    try:
        start_time = time.time()
        
        # Проверка авторизации (только для админов)
        user_type = check_auth(request)
        if not user_type or user_type != 'admin':
            return Response(status_code=403)
        
        # Если год не указан, используем текущий
        if not year:
            year = datetime.now().year
            
        # Получаем финансовые данные
        financial_data = get_annual_financial_data(db, year)
        yearly_totals = calculate_yearly_totals(financial_data)
        apartment_summary = get_apartment_summary(financial_data)
        
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = f"Дашборд {year}"
        
        # Стили
        header_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        header_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовки месяцев
        months = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь', 'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
        
        # Первая строка заголовков
        ws.merge_cells('A1:A2')
        ws['A1'] = 'Объект'
        ws['A1'].fill = header_fill
        ws['A1'].font = header_font
        ws['A1'].alignment = center_align
        ws['A1'].border = border
        
        # Итого
        ws.merge_cells('B1:D1')
        ws['B1'] = 'Итого'
        ws['B1'].fill = header_fill
        ws['B1'].font = header_font
        ws['B1'].alignment = center_align
        ws['B1'].border = border
        
        # Месяцы
        col_idx = 5
        for month_name in months:
            end_col_idx = col_idx + 2
            cell = ws.cell(row=1, column=col_idx)
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=end_col_idx)
            cell.value = month_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            # Apply border to merged cells
            for i in range(3):
                ws.cell(row=1, column=col_idx + i).border = border
            col_idx += 3
            
        # Вторая строка заголовков (Д/Р/П)
        headers = ['Доход', 'Расход', 'Прибыль']
        
        # Для Итого
        for i, header in enumerate(headers):
            cell = ws.cell(row=2, column=2 + i)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            
        # Для месяцев
        col_idx = 5
        for _ in range(12):
            for header in headers:
                cell = ws.cell(row=2, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border
                col_idx += 1
                
        # Данные
        row_idx = 3
        
        # Стили для данных
        text_success = Font(color="198754")  # Green
        text_danger = Font(color="DC3545")   # Red
        
        for base_apartment, apartment_data in apartment_summary.items():
            # Объект
            cell = ws.cell(row=row_idx, column=1)
            cell.value = base_apartment
            cell.font = Font(bold=True)
            cell.border = border
            
            # Итого по объекту
            # Доход
            cell = ws.cell(row=row_idx, column=2)
            cell.value = apartment_data['total_income']
            cell.number_format = '#,##0'
            cell.font = text_success
            cell.border = border
            
            # Расход
            cell = ws.cell(row=row_idx, column=3)
            cell.value = apartment_data['total_expenses']
            cell.number_format = '#,##0'
            cell.font = text_danger
            cell.border = border
            
            # Прибыль
            cell = ws.cell(row=row_idx, column=4)
            cell.value = apartment_data['total_profit']
            cell.number_format = '#,##0'
            cell.font = text_success if apartment_data['total_profit'] >= 0 else text_danger
            cell.border = border
            
            # По месяцам
            col_idx = 5
            for month in range(1, 13):
                month_key = f"{year}-{month:02d}"
                income = 0
                expenses = 0
                profit = 0
                
                if month_key in financial_data:
                    month_data = financial_data[month_key]
                    if base_apartment in month_data['objects']:
                        obj_data = month_data['objects'][base_apartment]
                        income = obj_data['income']
                        expenses = obj_data['expenses']
                        profit = obj_data['profit']
                
                # Доход
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = income if income != 0 else '-'
                if income != 0:
                    cell.number_format = '#,##0'
                    cell.font = text_success
                else:
                    cell.alignment = Alignment(horizontal='center')
                cell.border = border
                
                # Расход
                cell = ws.cell(row=row_idx, column=col_idx + 1)
                cell.value = expenses if expenses != 0 else '-'
                if expenses != 0:
                    cell.number_format = '#,##0'
                    cell.font = text_danger
                else:
                    cell.alignment = Alignment(horizontal='center')
                cell.border = border
                
                # Прибыль
                cell = ws.cell(row=row_idx, column=col_idx + 2)
                cell.value = profit if profit != 0 else '-'
                if profit != 0:
                    cell.number_format = '#,##0'
                    cell.font = text_success if profit >= 0 else text_danger
                else:
                    cell.alignment = Alignment(horizontal='center')
                cell.border = border
                
                col_idx += 3
            
            row_idx += 1
            
        # Общие расходы
        cell = ws.cell(row=row_idx, column=1)
        cell.value = "Общие расходы"
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid") # Warning color
        cell.border = border
        
        # Итого общие расходы
        cell = ws.cell(row=row_idx, column=2)
        cell.value = "-"
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        cell.border = border
        
        cell = ws.cell(row=row_idx, column=3)
        cell.value = yearly_totals['total_general_expenses']
        cell.number_format = '#,##0'
        cell.font = text_danger
        cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        cell.border = border
        
        cell = ws.cell(row=row_idx, column=4)
        cell.value = "-"
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        cell.border = border
        
        col_idx = 5
        for month in range(1, 13):
            month_key = f"{year}-{month:02d}"
            general_expenses = 0
            if month_key in financial_data:
                general_expenses = financial_data[month_key].get('general_expenses', 0)
                
            # Доход (-)
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = "-"
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            cell.border = border
            
            # Расход
            cell = ws.cell(row=row_idx, column=col_idx + 1)
            cell.value = general_expenses if general_expenses != 0 else '-'
            if general_expenses != 0:
                cell.number_format = '#,##0'
                cell.font = text_danger
            else:
                cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            cell.border = border
            
            # Прибыль (-)
            cell = ws.cell(row=row_idx, column=col_idx + 2)
            cell.value = "-"
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            cell.border = border
            
            col_idx += 3
            
        row_idx += 1
        
        # ИТОГО
        cell = ws.cell(row=row_idx, column=1)
        cell.value = "ИТОГО"
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid") # Secondary color
        cell.border = border
        
        # Итого всего
        # Доход
        cell = ws.cell(row=row_idx, column=2)
        cell.value = yearly_totals['total_income']
        cell.number_format = '#,##0'
        cell.font = text_success
        cell.fill = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")
        cell.border = border
        
        # Расход
        cell = ws.cell(row=row_idx, column=3)
        cell.value = yearly_totals['total_expenses'] + yearly_totals.get('total_general_expenses', 0)
        cell.number_format = '#,##0'
        cell.font = text_danger
        cell.fill = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")
        cell.border = border
        
        # Прибыль
        total_profit = yearly_totals['total_profit'] - yearly_totals.get('total_general_expenses', 0)
        cell = ws.cell(row=row_idx, column=4)
        cell.value = total_profit
        cell.number_format = '#,##0'
        cell.font = text_success if total_profit >= 0 else text_danger
        cell.fill = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")
        cell.border = border
        
        # По месяцам
        col_idx = 5
        for month in range(1, 13):
            month_key = f"{year}-{month:02d}"
            income = 0
            expenses = 0
            profit = 0
            general_expenses = 0
            
            if month_key in financial_data:
                month_data = financial_data[month_key]
                income = month_data['total_income']
                expenses = month_data['total_expenses']
                profit = month_data['total_profit']
                general_expenses = month_data.get('general_expenses', 0)
            
            # Доход
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = income
            cell.number_format = '#,##0'
            cell.font = text_success
            cell.fill = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")
            cell.border = border
            
            # Расход
            cell = ws.cell(row=row_idx, column=col_idx + 1)
            cell.value = expenses + general_expenses
            cell.number_format = '#,##0'
            cell.font = text_danger
            cell.fill = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")
            cell.border = border
            
            # Прибыль
            month_profit = profit - general_expenses
            cell = ws.cell(row=row_idx, column=col_idx + 2)
            cell.value = month_profit
            cell.number_format = '#,##0'
            cell.font = text_success if month_profit >= 0 else text_danger
            cell.fill = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")
            cell.border = border
            
            col_idx += 3
            
        # Автоширина колонок
        ws.column_dimensions['A'].width = 25
        for col in range(2, col_idx):
            col_letter = chr(64 + col) if col <= 26 else f'A{chr(64 + col - 26)}'
            ws.column_dimensions[col_letter].width = 15
            
        # Сохраняем в BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        filename = f"dashboard_{year}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        end_time = time.time()
        logger.info(f"Exported dashboard for {year} in {end_time - start_time:.2f} seconds")
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Error in export_dashboard_to_excel: {e}")
        return Response(status_code=500, content="Internal Server Error")
